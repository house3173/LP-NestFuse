import numpy as np
from PIL import Image
from Metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


def write_excel(excel_name, worksheet_name, column_index=0, data=None, start_column=1):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        # Nếu file không tồn tại, tạo Workbook mới
        workbook = Workbook()

    # Kiểm tra và lấy hoặc tạo sheet mới
    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    # Lấy tên cột từ chỉ số bắt đầu, cộng với index của cột
    column = get_column_letter(start_column + column_index)

    # Chèn dữ liệu vào cột chỉ định
    for i, value in enumerate(data):
        cell = worksheet[column + str(i + 1)]
        cell.value = value

    # Lưu workbook
    workbook.save(excel_name)


def evaluation_one(ir_name, vi_name, f_name):
    f_img = Image.open(f_name).convert('L')
    ir_img = Image.open(ir_name).convert('L')
    vi_img = Image.open(vi_name).convert('L')

    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)
    SD = SD_function(f_img_double)
    AG = AG_function(f_img_double)
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    return EN, MI, AG, SD, VIF,CC, Qabf


if __name__ == '__main__':
    with_mean = True
    dataroot = r'./datasets'
    results_root = './Results'
    dataset = 'TNO'
    ir_dir = os.path.join(dataroot, dataset, 'ir').replace("\\", "/")
    vi_dir = os.path.join(dataroot, dataset, 'vi').replace("\\", "/")
    f_dir = os.path.join(results_root, dataset).replace("\\", "/")
    save_dir = './Metric'
    os.makedirs(save_dir, exist_ok=True)
    sum_dir = './datasets/TNO/sum/'

    metric_save_name = os.path.join(save_dir, 'compare_{}.xlsx'.format(dataset)).replace("\\", "/")
    filelist = natsorted(os.listdir(ir_dir))

    Method_list = ['DenseFuse','CrossFuse', 'MCPFusion', 'BTSFusion', 'RFN-Nest', 'LP-NestFuse']
    # Method_list = ['Laplacian-Pyramid']
    for i, Method in enumerate(Method_list):
        EN_list = []
        MI_list = []
        AG_list = []
        SD_list = []
        CC_list = []
        VIF_list = []
        Qabf_list = []
        filename_list = ['']
        sub_f_dir = os.path.join(f_dir, Method).replace("\\", "/")
        eval_bar = tqdm(filelist)
        for _, item in enumerate(eval_bar):
            ir_name = os.path.join(ir_dir, '{:02}'.format(item)).replace("\\", "/")
            vi_name = os.path.join(vi_dir, '{:02}'.format(item)).replace("\\", "/")
            f_name = os.path.join(sub_f_dir, '{:02}'.format(item)).replace("\\", "/")
            EN, MI, AG, SD, VIF,CC, Qabf = evaluation_one(ir_name, vi_name, f_name)
            EN_list.append(EN)
            MI_list.append(MI)
            AG_list.append(AG)
            SD_list.append(SD)
            CC_list.append(CC)
            VIF_list.append(VIF)
            Qabf_list.append(Qabf)
            filename_list.append(item)
            eval_bar.set_description("{} | {}".format(Method, item))
        if with_mean:
            EN_list.append(np.mean(EN_list))
            MI_list.append(np.mean(MI_list))
            AG_list.append(np.mean(AG_list))
            SD_list.append(np.mean(SD_list))
            CC_list.append(np.mean(CC_list))
            VIF_list.append(np.mean(VIF_list))
            Qabf_list.append(np.mean(Qabf_list))
            filename_list.append('mean')

        EN_list = [round(x, 3) for x in EN_list]
        MI_list = [round(x, 3) for x in MI_list]
        AG_list = [round(x, 3) for x in AG_list]
        SD_list = [round(x, 3) for x in SD_list]
        CC_list = [round(x, 3) for x in CC_list]
        VIF_list = [round(x, 3) for x in VIF_list]
        Qabf_list = [round(x, 3) for x in Qabf_list]

        EN_list.insert(0, '{}'.format(Method))
        MI_list.insert(0, '{}'.format(Method))
        AG_list.insert(0, '{}'.format(Method))
        SD_list.insert(0, '{}'.format(Method))
        CC_list.insert(0, '{}'.format(Method))
        VIF_list.insert(0, '{}'.format(Method))
        Qabf_list.insert(0, '{}'.format(Method))
        if i == 0:
            write_excel(metric_save_name, 'EN', 0, filename_list)
            write_excel(metric_save_name, "MI", 0, filename_list)
            write_excel(metric_save_name, "AG", 0, filename_list)
            write_excel(metric_save_name, "SD", 0, filename_list)
            write_excel(metric_save_name, "CC", 0, filename_list)
            write_excel(metric_save_name, "VIF", 0, filename_list)
            write_excel(metric_save_name, "Qabf", 0, filename_list)
        write_excel(metric_save_name, 'EN', i + 1, EN_list,1)
        write_excel(metric_save_name, 'MI', i + 1, MI_list,1)
        write_excel(metric_save_name, 'AG', i + 1, AG_list,1)
        write_excel(metric_save_name, 'SD', i + 1, SD_list,1)
        write_excel(metric_save_name, 'CC', i + 1, CC_list,1)
        write_excel(metric_save_name, 'VIF', i + 1, VIF_list,1)
        write_excel(metric_save_name, 'Qabf', i + 1, Qabf_list,1)

